from __future__ import annotations

import csv
import io
import json
import mimetypes
from collections.abc import Iterable
from dataclasses import dataclass, field
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

SUPPORTED_EXTENSIONS = {
    ".txt",
    ".docx",
    ".csv",
    ".xlsx",
    ".pdf",
    ".jpg",
    ".jpeg",
    ".png",
    ".webp",
}
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp"}
IMAGE_CONTENT_MARKER = "The source is provided as an attached image."


class SourceLoadError(RuntimeError):
    pass


class UnsupportedSourceError(SourceLoadError):
    pass


@dataclass(frozen=True)
class SourceSection:
    key: str
    label: str
    text: str

    def render(self) -> str:
        return (
            f"[{self.label}]\n{self.text}".strip() if self.label else self.text.strip()
        )


@dataclass(frozen=True)
class ImageAttachment:
    label: str
    section_key: str
    mime_type: str
    data: bytes


@dataclass(frozen=True)
class SourceChunk:
    text: str
    attachments: tuple[ImageAttachment, ...]


@dataclass
class PreparedSource:
    path: Path
    source_name: str
    source_type: str
    processing_path: str
    sections: list[SourceSection]
    attachments: list[ImageAttachment] = field(default_factory=list)
    loader_metadata: dict[str, Any] = field(default_factory=dict)

    @property
    def content_text(self) -> str:
        return "\n\n".join(section.render() for section in self.sections)

    def needs_chunking(self, max_chars: int, max_images: int) -> bool:
        return len(self.content_text) > max_chars or len(self.attachments) > max_images

    def chunks(self, target_chars: int, max_images: int) -> list[SourceChunk]:
        if target_chars < 1:
            raise ValueError("target_chars must be positive")
        if max_images < 1:
            raise ValueError("max_images must be positive")

        attachment_map: dict[str, list[ImageAttachment]] = {}
        for attachment in self.attachments:
            attachment_map.setdefault(attachment.section_key, []).append(attachment)

        pieces: list[tuple[str, list[ImageAttachment]]] = []
        for section in self.sections:
            rendered_pieces = _split_section(section, target_chars)
            linked = attachment_map.get(section.key, [])
            for index, text in enumerate(rendered_pieces):
                pieces.append((text, linked if index == 0 else []))

        chunks: list[SourceChunk] = []
        current_text: list[str] = []
        current_images: list[ImageAttachment] = []
        current_chars = 0

        def flush() -> None:
            nonlocal current_text, current_images, current_chars
            if current_text:
                chunks.append(
                    SourceChunk(
                        text="\n\n".join(current_text),
                        attachments=tuple(current_images),
                    )
                )
            current_text = []
            current_images = []
            current_chars = 0

        for text, images in pieces:
            separator = 2 if current_text else 0
            exceeds_text = (
                current_text and current_chars + separator + len(text) > target_chars
            )
            exceeds_images = (
                current_text and len(current_images) + len(images) > max_images
            )
            if exceeds_text or exceeds_images:
                flush()
            current_text.append(text)
            current_images.extend(images)
            current_chars += (2 if current_chars else 0) + len(text)
        flush()
        return chunks


def _split_section(section: SourceSection, target_chars: int) -> list[str]:
    prefix = f"[{section.label}]\n" if section.label else ""
    available = max(1, target_chars - len(prefix))
    if len(section.text) <= available:
        return [prefix + section.text]

    pieces: list[str] = []
    current: list[str] = []
    current_size = 0
    for line in section.text.splitlines(keepends=True):
        if len(line) > available:
            if current:
                pieces.append(prefix + "".join(current).rstrip())
                current = []
                current_size = 0
            for start in range(0, len(line), available):
                pieces.append(prefix + line[start : start + available].rstrip())
            continue
        if current and current_size + len(line) > available:
            pieces.append(prefix + "".join(current).rstrip())
            current = []
            current_size = 0
        current.append(line)
        current_size += len(line)
    if current:
        pieces.append(prefix + "".join(current).rstrip())
    return pieces


def load_source(
    path: Path,
    *,
    pdf_min_text_chars: int = 80,
    pdf_image_heavy_text_chars: int = 400,
    pdf_image_area_ratio: float = 0.65,
    pdf_render_dpi: int = 144,
) -> PreparedSource:
    extension = path.suffix.lower()
    if extension not in SUPPORTED_EXTENSIONS:
        raise UnsupportedSourceError(
            f"Unsupported source format: {extension or '(none)'}"
        )
    if not path.is_file():
        raise SourceLoadError(f"Source is not a file: {path}")

    if extension == ".txt":
        return _load_txt(path)
    if extension == ".docx":
        return _load_docx(path)
    if extension == ".csv":
        return _load_csv(path)
    if extension == ".xlsx":
        return _load_xlsx(path)
    if extension == ".pdf":
        return _load_pdf(
            path,
            min_text_chars=pdf_min_text_chars,
            image_heavy_text_chars=pdf_image_heavy_text_chars,
            image_area_ratio=pdf_image_area_ratio,
            render_dpi=pdf_render_dpi,
        )
    return _load_image(path)


def _load_txt(path: Path) -> PreparedSource:
    text, encoding = _read_text(path)
    return PreparedSource(
        path=path,
        source_name=path.name,
        source_type="txt",
        processing_path="text",
        sections=[SourceSection("text", "Text file", text)],
        loader_metadata={"encoding": encoding},
    )


def _read_text(path: Path) -> tuple[str, str]:
    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "utf-16", "cp1252"):
        try:
            return raw.decode(encoding), encoding
        except UnicodeDecodeError:
            continue
    raise SourceLoadError(
        f"Could not decode text file without replacement: {path.name}"
    )


def _load_docx(path: Path) -> PreparedSource:
    try:
        from docx import Document
        from docx.table import Table
        from docx.text.paragraph import Paragraph
    except ImportError as exc:
        raise SourceLoadError("python-docx is required to parse DOCX files") from exc

    try:
        document = Document(path)
        sections: list[SourceSection] = []
        lines: list[str] = []
        label = "Document beginning"
        section_number = 1
        table_number = 0

        def flush() -> None:
            nonlocal lines, section_number
            if any(line.strip() for line in lines):
                sections.append(
                    SourceSection(
                        key=f"docx-section-{section_number}",
                        label=label,
                        text="\n".join(lines).strip(),
                    )
                )
                section_number += 1
            lines = []

        for child in document.element.body.iterchildren():
            if child.tag.endswith("}p"):
                paragraph = Paragraph(child, document)
                text = paragraph.text
                style_name = paragraph.style.name if paragraph.style is not None else ""
                if style_name.lower().startswith("heading") and text.strip():
                    flush()
                    label = f"{style_name}: {text.strip()}"
                    lines.append(text.strip())
                elif text.strip():
                    lines.append(text)
            elif child.tag.endswith("}tbl"):
                table_number += 1
                table = Table(child, document)
                lines.append(_serialize_docx_table(table, table_number))
        flush()
    except SourceLoadError:
        raise
    except Exception as exc:
        raise SourceLoadError(f"Failed to parse DOCX {path.name}: {exc}") from exc

    if not sections:
        raise SourceLoadError(
            f"DOCX has no extractable paragraphs or tables: {path.name}"
        )
    return PreparedSource(
        path=path,
        source_name=path.name,
        source_type="docx",
        processing_path="text",
        sections=sections,
        loader_metadata={"section_count": len(sections), "table_count": table_number},
    )


def _serialize_docx_table(table: Any, number: int) -> str:
    rows = [[cell.text for cell in row.cells] for row in table.rows]
    output = [f"[Table {number}]"]
    if rows:
        output.append("Columns: " + json.dumps(rows[0], ensure_ascii=False))
        for index, row in enumerate(rows[1:], start=1):
            output.append(f"Row {index}: " + json.dumps(row, ensure_ascii=False))
    return "\n".join(output)


def _load_csv(path: Path) -> PreparedSource:
    text, encoding = _read_text(path)
    try:
        dialect = csv.Sniffer().sniff(text[:8192], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    rows = list(csv.reader(io.StringIO(text, newline=""), dialect))
    rows = _unwrap_single_column_csv(rows)
    if not rows:
        raise SourceLoadError(f"CSV is empty: {path.name}")

    columns = rows[0]
    sections: list[SourceSection] = []
    data_rows = rows[1:]
    if not data_rows:
        sections.append(
            SourceSection("csv-rows-1", "CSV rows 1-1", _serialize_rows(columns, [], 1))
        )
    else:
        for group_index, start in enumerate(range(0, len(data_rows), 100), start=1):
            group = data_rows[start : start + 100]
            first_row = start + 2
            last_row = first_row + len(group) - 1
            sections.append(
                SourceSection(
                    key=f"csv-rows-{group_index}",
                    label=f"CSV rows {first_row}-{last_row}",
                    text=_serialize_rows(columns, group, first_row),
                )
            )
    return PreparedSource(
        path=path,
        source_name=path.name,
        source_type="csv",
        processing_path="text",
        sections=sections,
        loader_metadata={
            "encoding": encoding,
            "column_count": len(columns),
            "data_row_count": len(data_rows),
        },
    )


def _unwrap_single_column_csv(rows: list[list[str]]) -> list[list[str]]:
    if not rows or any(len(row) != 1 for row in rows):
        return rows
    candidates = rows[0][0]
    if not any(delimiter in candidates for delimiter in (",", ";", "\t", "|")):
        return rows
    try:
        dialect = csv.Sniffer().sniff(candidates, delimiters=",;\t|")
        reparsed = [next(csv.reader([row[0]], dialect)) for row in rows]
    except (csv.Error, StopIteration):
        return rows
    return reparsed if max((len(row) for row in reparsed), default=0) > 1 else rows


def _serialize_rows(columns: list[Any], rows: list[list[Any]], first_row: int) -> str:
    output = ["Columns: " + json.dumps(columns, ensure_ascii=False, default=str)]
    for offset, row in enumerate(rows):
        output.append(
            f"Row {first_row + offset}: "
            + json.dumps(row, ensure_ascii=False, default=str)
        )
    return "\n".join(output)


def _load_xlsx(path: Path) -> PreparedSource:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SourceLoadError("openpyxl is required to parse XLSX files") from exc

    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
        sections: list[SourceSection] = []
        sheet_metadata: list[dict[str, Any]] = []
        try:
            for sheet_index, worksheet in enumerate(workbook.worksheets, start=1):
                rows: list[list[Any]] = []
                for values in worksheet.iter_rows(values_only=True):
                    row = [_xlsx_value(value) for value in values]
                    while row and row[-1] is None:
                        row.pop()
                    if row and any(value is not None for value in row):
                        rows.append(row)
                sheet_metadata.append(
                    {"name": worksheet.title, "nonempty_rows": len(rows)}
                )
                if not rows:
                    sections.append(
                        SourceSection(
                            key=f"xlsx-sheet-{sheet_index}-rows-0",
                            label=f"Sheet: {worksheet.title}",
                            text="(empty sheet)",
                        )
                    )
                    continue
                columns = rows[0]
                data_rows = rows[1:]
                groups = list(_groups(data_rows, 200)) or [[]]
                for group_index, group in enumerate(groups, start=1):
                    first_row = (group_index - 1) * 200 + 2
                    last_row = first_row + max(0, len(group) - 1)
                    label = f"Sheet: {worksheet.title}; rows {first_row}-{last_row}"
                    sections.append(
                        SourceSection(
                            key=f"xlsx-sheet-{sheet_index}-rows-{group_index}",
                            label=label,
                            text=_serialize_rows(columns, group, first_row),
                        )
                    )
        finally:
            workbook.close()
    except Exception as exc:
        raise SourceLoadError(f"Failed to parse XLSX {path.name}: {exc}") from exc

    return PreparedSource(
        path=path,
        source_name=path.name,
        source_type="xlsx",
        processing_path="text",
        sections=sections,
        loader_metadata={"sheets": sheet_metadata},
    )


def _xlsx_value(value: Any) -> Any:
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return value


def _groups(values: list[Any], size: int) -> Iterable[list[Any]]:
    for start in range(0, len(values), size):
        yield values[start : start + size]


def _load_image(path: Path) -> PreparedSource:
    try:
        from PIL import Image
    except ImportError as exc:
        raise SourceLoadError("Pillow is required to validate image files") from exc

    data = path.read_bytes()
    try:
        with Image.open(io.BytesIO(data)) as image:
            image.verify()
            detected_format = (image.format or "").upper()
    except Exception as exc:
        raise SourceLoadError(f"Invalid image {path.name}: {exc}") from exc

    format_to_mime = {
        "JPEG": "image/jpeg",
        "PNG": "image/png",
        "WEBP": "image/webp",
    }
    mime_type = (
        format_to_mime.get(detected_format) or mimetypes.guess_type(path.name)[0]
    )
    if mime_type not in {"image/jpeg", "image/png", "image/webp"}:
        raise SourceLoadError(
            f"Unsupported image encoding in {path.name}: {detected_format}"
        )
    section = SourceSection("original-image", "", IMAGE_CONTENT_MARKER)
    return PreparedSource(
        path=path,
        source_name=path.name,
        source_type=path.suffix.lower().lstrip("."),
        processing_path="vision",
        sections=[section],
        attachments=[ImageAttachment(path.name, section.key, mime_type, data)],
        loader_metadata={"detected_format": detected_format},
    )


def _load_pdf(
    path: Path,
    *,
    min_text_chars: int,
    image_heavy_text_chars: int,
    image_area_ratio: float,
    render_dpi: int,
) -> PreparedSource:
    try:
        import pymupdf as fitz
    except ImportError as exc:
        raise SourceLoadError("PyMuPDF is required to parse PDF files") from exc

    sections: list[SourceSection] = []
    attachments: list[ImageAttachment] = []
    page_metadata: list[dict[str, Any]] = []
    try:
        document = fitz.open(path)
        try:
            if document.needs_pass:
                raise SourceLoadError(f"Encrypted PDF requires a password: {path.name}")
            scale = render_dpi / 72.0
            for page_index, page in enumerate(document, start=1):
                text = page.get_text("text", sort=True).strip()
                meaningful_chars = sum(character.isalnum() for character in text)
                page_area = max(float(page.rect.get_area()), 1.0)
                image_area = 0.0
                for image in page.get_image_info():
                    try:
                        image_area += float(fitz.Rect(image["bbox"]).get_area())
                    except (KeyError, TypeError, ValueError):
                        continue
                image_ratio = min(image_area / page_area, 1.0)
                use_vision = meaningful_chars < min_text_chars or (
                    image_ratio >= image_area_ratio
                    and meaningful_chars < image_heavy_text_chars
                )
                section_key = f"pdf-page-{page_index}"
                if use_vision:
                    pixmap = page.get_pixmap(
                        matrix=fitz.Matrix(scale, scale),
                        alpha=False,
                        colorspace=fitz.csRGB,
                    )
                    attachments.append(
                        ImageAttachment(
                            label=f"{path.name} — page {page_index}",
                            section_key=section_key,
                            mime_type="image/png",
                            data=pixmap.tobytes("png"),
                        )
                    )
                    page_text = (
                        "This page is provided as an attached rendered image because "
                        "extractable text was insufficient."
                    )
                    mode = "vision"
                else:
                    page_text = text
                    mode = "text"
                sections.append(
                    SourceSection(section_key, f"Page {page_index}", page_text)
                )
                page_metadata.append(
                    {
                        "page": page_index,
                        "mode": mode,
                        "meaningful_text_chars": meaningful_chars,
                        "image_area_ratio": round(image_ratio, 4),
                    }
                )
        finally:
            document.close()
    except SourceLoadError:
        raise
    except Exception as exc:
        raise SourceLoadError(f"Failed to parse PDF {path.name}: {exc}") from exc

    if not sections:
        raise SourceLoadError(f"PDF contains no pages: {path.name}")
    vision_pages = len(attachments)
    if vision_pages == 0:
        processing_path = "text"
    elif vision_pages == len(sections):
        processing_path = "vision"
    else:
        processing_path = "hybrid"
    return PreparedSource(
        path=path,
        source_name=path.name,
        source_type="pdf",
        processing_path=processing_path,
        sections=sections,
        attachments=attachments,
        loader_metadata={
            "page_count": len(sections),
            "rendered_page_count": vision_pages,
            "pages": page_metadata,
        },
    )
